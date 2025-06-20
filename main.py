from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium import webdriver
from urllib.parse import quote
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime
import requests
import time
import re

wb = load_workbook("album.xlsx")
ws = wb.active
url_base = 'https://www.ligapokemon.com.br/?view=cards/card&card='

options = Options()
options.add_argument('--incognito')
#options.add_argument("--headless=new")
options.add_argument("--no-sandbox")
driver = webdriver.Chrome(options=options)


driver.delete_all_cookies()
driver.get("Https:/www.google.com")
driver.execute_script("window.localStorage.clear();")


driver.maximize_window()
wait = WebDriverWait(driver, 10)

for i, row in enumerate(ws.iter_rows(min_row=2)):
    print('_____________________________________________________________________________________________\n')
    dados = {
        "nome": row[0].value,
        "numero": row[1].value,
        "condicao": row[2].value,
    }

    if None in dados.values():
        print("Informações incompletas, indo para a próxima...\n")
        continue
    else:
        nome = str(dados["nome"]).strip()
        numero = str(dados["numero"]).strip()
        condicao = str(dados["condicao"]).strip()


    nome_carta = f'{nome.strip()} ({numero.strip()})'
    print(f'Nome: {nome_carta} condição: {condicao}')

    url = f'{url_base}{quote(nome_carta, safe="()/'")}'
    print(f'URL acessada: {url}')
    driver.get(url)

    # - Marca o checkbox do estado NM
    map_de_condicao = {
    'M': 'filter_4_0',   # M Nova
    'NM': 'filter_4_1',  # NM Praticamente Nova ou superior
    'SP': 'filter_4_2',  # SP Usada Levemente ou superior
    'MP': 'filter_4_3',  # MP Usada Moderadamente ou superior
    'HP': 'filter_4_4',  # HP Muito Usada ou superior
    'D': 'filter_4_5'    # D Danificada ou superior
    }

    try:
        # Checkbox de condicao
        try:
            limpar_filtros = wait.until(EC.presence_of_element_located((By.ID, 'group-4')))
            
            try:
                checkbox_nm = wait.until(EC.element_to_be_clickable((By.ID, map_de_condicao[condicao])))
                
                if checkbox_nm.is_displayed():
                    driver.execute_script("arguments[0].scrollIntoView(true);", checkbox_nm)
                    
                    if limpar_filtros.find_element(By.ID, 'clear-4').is_displayed():
                        if 'filtro-opcao-vazia' not in checkbox_nm.get_attribute('class'):
                            pass
                        else:
                            limpar_btn = wait.until(EC.element_to_be_clickable((By.ID, 'clear-4')))
                            driver.execute_script("arguments[0].scrollIntoView(true);", limpar_btn)
                            limpar_btn.click()
                            wait.until(EC.invisibility_of_element_located((By.ID, 'layer-loading')))
                            
                            # Tentar novamente após limpar
                            checkbox_nm = wait.until(EC.element_to_be_clickable((By.ID, map_de_condicao[condicao])))
                            driver.execute_script("arguments[0].scrollIntoView(true);", checkbox_nm)
                            checkbox_nm.click()
                            wait.until(EC.invisibility_of_element_located((By.ID, 'layer-loading')))
                    else:
                        driver.execute_script("arguments[0].scrollIntoView(true);", checkbox_nm)
                        checkbox_nm.click()
                        wait.until(EC.invisibility_of_element_located((By.ID, 'layer-loading')))
            except Exception as e:
                print('Não existem cartas à venda nesse filtro, passando para a próxima...')
                continue
        except Exception as e:
            print('Não existem cartas à venda nesse filtro, passando para a próxima...')
            continue
                
        except Exception as e:
            print(f'Falha geral no filtro de checkbox, erro: {e}')
            continue

        except Exception as e:
            print(f'Falha no filtro de checkbox, erro: {e}')
            continue

        try:
            img_element = wait.until(EC.presence_of_element_located((By.ID, "featuredImage")))
            img_url = img_element.get_attribute("src")

            if img_url.startswith("//"):
                img_url = "https:" + img_url
            
            nome_arquivo = re.sub(r'[^\w\-_.()]', '_', nome_carta)
            img_path = f"imagens/{nome_arquivo}.jpg"

            img_data = requests.get(img_url).content
            with open(img_path, 'wb') as f:
                f.write(img_data)

            ws._images = [img for img in ws._images if img.anchor != f'E{i+2}']
            img_excel = ExcelImage(img_path)
            ws.add_image(img_excel, f"E{i + 2}")
        except Exception as e:
            print(f'Falha na alocação da imagem, erro: {e}')


        # Botão de comprar 
        marketplace_stores = driver.find_element(By.ID, 'marketplace-stores')
        div = marketplace_stores.find_element(By.XPATH, './div[1]')
        if div.is_displayed():
            pass
        else:
            try:
                i = 2
                while div.is_displayed() == 0:
                    div = marketplace_stores.find_element(By.XPATH, f'./div[{i}]')
                    if div.is_displayed():
                        break
                    i += 1
            except Exception as e:
                print(f'Falha no botão de compra, erro: {e}')

        botao_container = div.find_element(By.CLASS_NAME, "buttons")
        botao = botao_container.find_element(By.XPATH, './div[1]')
        driver.execute_script("arguments[0].click();", botao)
        time.sleep(0.3)

        # Ir pro carrinho e pegar preço
        try:
            driver.get('https://www.ligapokemon.com.br/?view=mp/carrinho')
            preco = wait.until(EC.visibility_of_element_located((By.ID, 'sum_preco_0'))).text
            preco = preco[3:].replace(',', '.').strip()
            print(f'Preco: {preco}')
        except Exception as e:
            print(f'Falha na troca de págino ou seleção de preço, erro: {e}')
            continue


        # Tirar do carrinho e aceitar alerta
        try:
            tirar_carrinho = driver.find_element(By.CSS_SELECTOR, '[title="Remover do Carrinho"]')
            onclick = tirar_carrinho.get_attribute("onclick")
            driver.execute_script(onclick)

            wait = WebDriverWait(driver, timeout=2)
            alert = wait.until(lambda d : d.switch_to.alert)
            text = alert.text
            alert.accept()
            time.sleep(0.1)
        except Exception as e:
            print(f'Falha em tirar carta do carrinho ou confirmação de alerta, erro: {e}')
            continue

        row[3].value = float(preco)
        
    except Exception as e:
        print(e)

precos = [row[3].value for row in ws.iter_rows(min_row=2) if isinstance(row[3].value, (int, float))]
ws["F2"].value = sum(precos)
ws["G2"].value = datetime.now().strftime("%d/%m/%Y") 

wb.save("album.xlsx")
driver.quit()